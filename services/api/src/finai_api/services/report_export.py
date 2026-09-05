"""Excel presentation of an immutable calculation; no Excel calculation authority."""

from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


def operating_workbook(result: dict[str, Any]) -> bytes:
    workbook = Workbook()
    workbook.remove(workbook.active)
    tables = {
        "Operating P&L": [
            ["REFERENCE RECONSTRUCTION — NOT APPROVED PETROLEUM ACTUALS"],
            ["Report line", "Calculated amount", "Availability", "Exact decimal amount"],
            *[
                [
                    m["id"].replace(".", " / "),
                    float(m["amount"]) if m["amount"] is not None else None,
                    m["state"],
                    m["amount"] or "Unavailable",
                ]
                for m in result["metrics"]
            ],
        ],
        "Source lineage": [
            ["SOURCE VALUES AND PROPOSED CLASSIFICATION"],
            ["Metric", "Sheet", "Row", "Source label", "Exact amount", "Source cells"],
            *[
                [
                    f["metric"],
                    f["source_sheet"],
                    f["source_row"],
                    f["label"],
                    f["amount"],
                    ", ".join(f["coordinates"]),
                ]
                for f in result["facts"]
            ],
        ],
        "Comparison": [
            ["DIFFERENCES ARE PRESERVED FOR REVIEW"],
            ["Metric", "Workbook cell", "Cached amount", "Calculated amount", "Difference"],
            *[
                [c["metric"], c["coordinate"], c["legacy_cached"], c["calculated"], c["difference"]]
                for c in result["comparisons"]
            ],
        ],
        "Method & context": [
            ["RECONSTRUCTION CONTEXT"],
            ["Property", "Value"],
            ["State", result["state"]],
            ["Source filename", result.get("filename", "")],
            ["Source SHA256", result["source_sha256"]],
            ["Calculation", result.get("calculation_id", "")],
            ["Contract", result["contract_version"]],
            ["Meaning", result["explanation"]],
            [
                "Units",
                "Source amount scale; no currency conversion. "
                "Source company/period/currency require review.",
            ],
            *[["Missing requirement", reason] for reason in result["missing_requirements"]],
        ],
    }
    for name, rows in tables.items():
        sheet = workbook.create_sheet(name)
        for row in rows:
            sheet.append(row)
        for row in sheet:
            for cell in row:
                # Never turn untrusted labels into executable spreadsheet formulas.
                if isinstance(cell.value, str):
                    cell.data_type = "s"
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                if cell.row <= 2:
                    cell.fill = PatternFill("solid", fgColor="173B49")
                    cell.font = Font(color="FFFFFF", bold=True)
        sheet.row_dimensions[1].height = 38
        sheet.row_dimensions[2].height = 30
        sheet.freeze_panes = "B3"
        sheet.auto_filter.ref = f"A2:{sheet.cell(sheet.max_row, sheet.max_column).coordinate}"
        sheet.sheet_view.showGridLines = False
        for column in range(1, sheet.max_column + 1):
            sheet.column_dimensions[sheet.cell(2, column).column_letter].width = (
                36 if column == 1 else 29
            )
        sheet.print_title_rows = "1:2"
        sheet.sheet_properties.pageSetUpPr.fitToPage = True
        sheet.page_setup.orientation = "landscape"
        sheet.page_setup.paperSize = sheet.PAPERSIZE_A4
        sheet.page_setup.fitToWidth = 1
        sheet.page_setup.fitToHeight = 0
    for row in workbook["Operating P&L"].iter_rows(min_row=3):
        row[1].number_format = '#,##0.00;[Red](#,##0.00);"—"'
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()
